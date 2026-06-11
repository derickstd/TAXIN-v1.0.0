"""
Export utilities for Excel and PDF generation.
"""
from io import BytesIO
from django.http import HttpResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors


def export_to_excel(filename, columns, data_rows, title=None):
    """
    Generate and return an Excel file.
    
    Args:
        filename (str): Output filename without extension (e.g., 'invoices')
        columns (list): List of column header names
        data_rows (list): List of lists/tuples of row data
        title (str): Optional title row
    
    Returns:
        HttpResponse: Excel file response
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    
    # Add title if provided
    if title:
        ws.merge_cells(f'A1:{chr(64 + len(columns))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1A3C6E", end_color="1A3C6E", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        start_row = 3
    else:
        start_row = 1
    
    # Add headers
    header_fill = PatternFill(start_color="E0F9FC", end_color="E0F9FC", fill_type="solid")
    header_font = Font(bold=True, color="0E7490")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.row_dimensions[start_row].height = 20
    
    # Add data rows
    for row_idx, row_data in enumerate(data_rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Adjust column widths
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = max(15, len(str(col_name)) + 2)
    
    # Generate response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


def export_to_pdf(filename, columns, data_rows, title=None, page_size=letter):
    """
    Generate and return a PDF file.
    
    Args:
        filename (str): Output filename without extension
        columns (list): List of column header names
        data_rows (list): List of lists/tuples of row data
        title (str): Optional title
        page_size: reportlab page size (letter or A4)
    
    Returns:
        HttpResponse: PDF file response
    """
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch
    )
    
    elements = []
    
    # Add title
    if title:
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1A3C6E'),
            spaceAfter=12,
            alignment=1  # center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2*inch))
    
    # Prepare table data
    table_data = [columns]
    table_data.extend(data_rows)
    
    # Create table
    table = Table(table_data, repeatRows=1)
    
    # Style table
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0E7490')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FDFF')]),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#A5F3FC')),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Add export info footer
    styles = getSampleStyleSheet()
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        spaceAfter=0,
        alignment=2  # right
    )
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(
        f"Exported on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Taxman256",
        footer_style
    ))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


def paginate_list(request, queryset, per_page=20):
    """
    Simple pagination helper.
    
    Args:
        request: Django request object
        queryset: Django QuerySet
        per_page: Items per page (default 20)
    
    Returns:
        tuple: (page_obj, is_paginated, has_more_than_20)
    """
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    paginator = Paginator(queryset, per_page)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    is_paginated = paginator.num_pages > 1
    has_more_than_20 = queryset.count() > 20
    
    return page_obj, is_paginated, has_more_than_20
